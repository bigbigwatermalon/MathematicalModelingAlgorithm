import math, random
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

class Population:
    # 种群的设计
    def __init__(self, size, chrom_size, cp, mp, gen_max, data):
        # 种群信息合
        self.x = []
        self.y = []
        self.individuals = []  # 个体集合
        self.fitness = []  # 个体适应度集
        self.selector_probability = []  # 个体选择概率集合
        self.new_individuals = []  # 新一代个体集合

        self.elitist = {'chromosome': [0, 0], 'fitness': 0, 'age': 0}  # 最佳个体的信息

        self.size = size  # 种群所包含的个体数
        self.chromosome_size = chrom_size  # 个体的染色体长度
        self.crossover_probability = cp  # 个体之间的交叉概率
        self.mutation_probability = mp  # 个体之间的变异概率

        self.generation_max = gen_max  # 种群进化的最大世代数
        self.age = 0  # 种群当前所处世代
        self.acess_data = data #######################################
        # 随机产生初始个体集，并将新一代个体、适应度、选择概率等集合以 0 值进行初始化
        v = 100
        for i in range(self.size):
            vx = ['0b' + self.create_bin(v) + self.create_bin(v) + self.create_bin(v),
                  '0b' + self.create_bin(v) + self.create_bin(v) + self.create_bin(v)]
            print(vx[0], vx[1])
            vx = [int(vx[0], 2), int(vx[1], 2)]
            self.individuals.append(vx)
            self.new_individuals.append([0, 0])
            self.fitness.append(0)
            self.selector_probability.append(0)
    def create_bin(self, v):
        x = bin(random.randint(0, v))
        x = x[2:]
        x = '0'*(7-len(x)) + x
        return x
    # 基于轮盘赌博机的选择
    def decode(self, interval, chromosome):           ###################此处
        '''将一个染色体 chromosome 映射为区间 interval 之内的数值'''
        d = interval[1] - interval[0]
        n = float(2 ** self.chromosome_size - 1)
        return (interval[0] + chromosome * d / n)

    def fitness_func(self, chrom1, chrom2):            ####################此处
        '''适应度函数，可以根据个体的两个染色体计算出该个体的适应度'''
        vx = [bin(chrom1)[2:], bin(chrom2)[2:]]
        #print(vx)
        try:
            x1 = int(int(vx[0][:-14], 2) * 101 / 127)
        except:
            x1 = 0
        try:
            y1 = int(int(vx[1][:-14], 2) * 101 / 127)
        except:
            y1 = 0
        try:
            x2 = int(int(vx[0][-14:-7], 2) * 101 / 127)
        except:
            x2 = 0
        try:
            y2 = int(int(vx[1][-14:-7], 2) * 101 / 127)
        except:
            y2 = 0
        (x3, y3) = (int(int(vx[0][-7:], 2) * 101 / 127), int(int(vx[1][-7:], 2) * 101 / 127))
        X = np.array([x1, y1, x2, y2, x3, y3]).reshape(3, 2)
        data_pre = self.fx(X)
        fitness = 1 / self.MS_error(data_pre, self.acess_data)
        print('age = %d fitness = %f (x1, y1) = (%d, %d) (x2, y2) = (%d, %d) (x3, y3) = (%d, %d) '%(self.age, fitness, X[0][0], X[0][1], X[1][0], X[1][1], X[2][0], X[2][1]))
        return fitness

    def single_fx(self, i, j, x):
        a = 0.05                                ##########那个不知道的是啥的值
        return 0.77 * math.exp(-(pow(a * (x[0] - i), 2) + pow(a * (x[1] - j), 2)))

    def fx(self, X):
        ##################计算浓度分布
        data_predict = np.zeros((101, 101))
        for x in X:
            for i in range(101):
                for j in range(101):
                    data_predict[i][j] += self.single_fx(i + 1, j + 1, x)
        #print(data.sum())
        #print('bb', data_predict.sum())
        #print(np.max(data_predict))
        return data_predict

    def MS_error(self, data, data_predict):  ###计算均方误差
        # print(data_predict.sum())
        error = 0
        # data = data * (data_predict.sum() / data.sum())
        #print('aa', data.sum())
        #print(data_predict.sum())
        for i in range(101):
            for j in range(101):
                error_x = pow(data[i][j] - data_predict[i][j], 2)
                error += error_x
        return error                #均方误差最小时，这个最大

    def evaluate(self):
        '''用于评估种群中的个体集合 self.individuals 中各个个体的适应度'''
        sp = self.selector_probability
        for i in range(self.size):
            self.fitness[i] = self.fitness_func(self.individuals[i][0],  # 将计算结果保存在 self.fitness 列表中
                                                self.individuals[i][1])
        ft_sum = sum(self.fitness)
        for i in range(self.size):
            sp[i] = self.fitness[i] / float(ft_sum)  # 得到各个个体的生存概率
        for i in range(1, self.size):
            sp[i] = sp[i] + sp[i - 1]  # 需要将个体的生存概率进行叠加，从而计算出各个个体的选择概率

    # 轮盘赌博机（选择）
    def select(self):
        (t, i) = (random.random(), 0)         ############？？？？？？？？？
        for p in self.selector_probability:
            if p > t:
                break
            i = i + 1
        return i

    # 交叉
    def cross(self, chrom1, chrom2):
        p = random.random()  # 随机概率
        n = 2 ** self.chromosome_size - 1
        if chrom1 != chrom2 and p < self.crossover_probability:
            t = random.randint(1, self.chromosome_size - 1)  # 随机选择一点（单点交叉）
            mask = n << t  # << 左移运算符
            (r1, r2) = (chrom1 & mask, chrom2 & mask)  # & 按位与运算符：参与运算的两个值,如果两个相应位都为1,则该位的结果为1,否则为0
            mask = n >> (self.chromosome_size - t)
            (l1, l2) = (chrom1 & mask, chrom2 & mask)
            (chrom1, chrom2) = (r1 + l2, r2 + l1)
        return (chrom1, chrom2)

    # 变异
    def mutate(self, chrom):
        p = random.random()
        if p < self.mutation_probability:
            t = random.randint(1, self.chromosome_size)
            mask1 = 1 << (t - 1)
            mask2 = chrom & mask1
            if mask2 > 0:
                chrom = chrom & (~mask2)  # ~ 按位取反运算符：对数据的每个二进制位取反,即把1变为0,把0变为1
            else:
                chrom = chrom ^ mask1  # ^ 按位异或运算符：当两对应的二进位相异时，结果为1
        return chrom

    # 保留最佳个体
    def reproduct_elitist(self):                    ##############这里
        # 与当前种群进行适应度比较，更新最佳个体
        j = -1
        for i in range(self.size):
            if self.elitist['fitness'] < self.fitness[i]:
                j = i
                self.elitist['fitness'] = self.fitness[i]
        if (j >= 0):
            self.elitist['chromosome'][0] = self.individuals[j][0]
            self.elitist['chromosome'][1] = self.individuals[j][1]
            self.elitist['age'] = self.age

    # 进化过程
    def evolve(self):
        indvs = self.individuals
        new_indvs = self.new_individuals
        # 计算适应度及选择概率
        self.evaluate()
        # 进化操作
        i = 0
        while True:
            # 选择两个个体，进行交叉与变异，产生新的种群
            idv1 = self.select()
            idv2 = self.select()
            # 交叉
            (idv1_x, idv1_y) = (indvs[idv1][0], indvs[idv1][1])
            (idv2_x, idv2_y) = (indvs[idv2][0], indvs[idv2][1])
            (idv1_x, idv2_x) = self.cross(idv1_x, idv2_x)
            (idv1_y, idv2_y) = self.cross(idv1_y, idv2_y)
            # 变异
            (idv1_x, idv1_y) = (self.mutate(idv1_x), self.mutate(idv1_y))
            (idv2_x, idv2_y) = (self.mutate(idv2_x), self.mutate(idv2_y))
            (new_indvs[i][0], new_indvs[i][1]) = (idv1_x, idv1_y)  # 将计算结果保存于新的个体集合self.new_individuals中
            (new_indvs[i + 1][0], new_indvs[i + 1][1]) = (idv2_x, idv2_y)
            # 判断进化过程是否结束
            i = i + 2  # 循环self.size/2次，每次从self.individuals 中选出2个
            if i >= self.size:
                break

        # 最佳个体保留
        # 如果在选择之前保留当前最佳个体，最终能收敛到全局最优解。
        self.reproduct_elitist()

        # 更新换代：用种群进化生成的新个体集合 self.new_individuals 替换当前个体集合
        for i in range(self.size):
            self.individuals[i][0] = self.new_individuals[i][0]
            self.individuals[i][1] = self.new_individuals[i][1]

    def run(self):
        '''根据种群最大进化世代数设定了一个循环。
        在循环过程中，调用 evolve 函数进行种群进化计算，并输出种群的每一代的个体适应度最大值、平均值和最小值。'''
        for i in range(self.generation_max):
            self.evolve()
            print('age = %d (x, y) = (%d, %d) fitness = %f' % (
            self.age, self.elitist['chromosome'][0], self.elitist['chromosome'][1], max(self.fitness)))
            print(i, max(self.fitness), sum(self.fitness) / self.size, min(self.fitness))
            self.x.append(self.age)
            self.y.append(max(self.fitness))
            self.age += 1
        print('age = %d (x, y) = (%d, %d)'%(self.elitist['age'], self.elitist['chromosome'][0], self.elitist['chromosome'][1]))
        plt.plot(self.x, self.y)
        plt.xlabel('age')
        plt.ylabel('fitness')
        plt.show()

if __name__ == '__main__':
    wb = load_workbook('Data01.xlsx')
    ws = wb['Sheet1']
    data = []
    for i in range(3, 104):
        line = ws[i]
        for element in line[2:-1]:
            data.append(element.value)
    data = np.array(data).reshape(101, -1)

    # 种群的个体数量为 100，染色体长度为 21，交叉概率为 0.8，变异概率为 0.1,进化最大世代数为 150
    pop = Population(100, 21, 0.8, 0.1, 150, data)
    pop.run()